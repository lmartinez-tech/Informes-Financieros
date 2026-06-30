from __future__ import annotations

from copy import copy
from datetime import datetime
from io import BytesIO
from pathlib import Path
import re
from typing import BinaryIO

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, PatternFill, Protection
from openpyxl.utils import get_column_letter

from utils.tlg_data_cleaning import MONTHS, load_tlg_trial_balance, normalize_text
from utils.tlg_financial_statements import prepare_tlg_detail


TEMPLATE_PATH = Path(__file__).resolve().parents[1] / "Informes_mensualizados_template.xlsx"
SUPPORTED_YEARS = {2024, 2025, 2026}


def _period_from_metadata(metadata: dict[str, str | None]) -> tuple[int, int]:
    month_name = str(metadata.get("mes") or "").strip().lower()
    year_text = str(metadata.get("anio") or "").strip()
    month = MONTHS.get(month_name)
    if month is None or not year_text.isdigit():
        raise ValueError(
            "No fue posible identificar el mes y el año del balance. "
            "Verifica que el encabezado indique el periodo."
        )
    return int(year_text), month


def _account4_values(
    detail: pd.DataFrame,
    result_account_code: str,
) -> tuple[dict[str, float], dict[str, float]]:
    data = detail.copy()
    data["CUENTA_4"] = (
        data["CODIGO_CUENTA"]
        .fillna("")
        .astype(str)
        .str.replace(r"\.0$", "", regex=True)
        .str.replace(r"\D", "", regex=True)
        .str[:4]
    )
    data = data[data["CUENTA_4"].str.len() == 4].copy()

    balance = data.groupby("CUENTA_4")["SALDO_FINAL"].sum().to_dict()
    balance[result_account_code] = float(
        data.loc[data["CLASE"].isin(["4", "5", "6", "7"]), "SALDO_FINAL"].sum()
    )
    data["VALOR_PYG"] = data["MOVIMIENTO_DEBITO"] - data["MOVIMIENTO_CREDITO"]
    income = data["CLASE"] == "4"
    data.loc[income, "VALOR_PYG"] = (
        data.loc[income, "MOVIMIENTO_CREDITO"]
        - data.loc[income, "MOVIMIENTO_DEBITO"]
    )
    pyg = data.groupby("CUENTA_4")["VALOR_PYG"].sum().to_dict()
    return balance, pyg


def _find_bce_month_column(worksheet, year: int, month: int) -> int:
    for column in range(1, min(worksheet.max_column, 48) + 1):
        value = worksheet.cell(2, column).value
        if isinstance(value, (datetime, pd.Timestamp)):
            if value.year == year and value.month == month:
                return column
        if month == 12 and isinstance(value, str):
            normalized = value.lower().replace(" ", "")
            if normalized.startswith("saldofinal") and str(year) in normalized:
                return column
    raise ValueError(
        f"La plantilla no contiene el periodo {month:02d}/{year} en la hoja BCE."
    )


def _find_pyg_month_column(worksheet, bce_column: int, year: int, month: int) -> int:
    bce_letter = get_column_letter(bce_column)
    reference_pattern = re.compile(
        rf"BCE!\$?{re.escape(bce_letter)}\$?2\b",
        flags=re.IGNORECASE,
    )
    for column in range(1, worksheet.max_column + 1):
        for row in range(1, min(worksheet.max_row, 6) + 1):
            value = worksheet.cell(row, column).value
            if isinstance(value, str) and reference_pattern.search(value):
                return column
            if isinstance(value, (datetime, pd.Timestamp)):
                if value.year == year and value.month == month:
                    return column
    raise ValueError(
        f"La plantilla no contiene el periodo {month:02d}/{year} en la hoja P Y G."
    )


def _find_pyg_total_column(worksheet, year: int) -> int:
    for row in range(1, min(worksheet.max_row, 8) + 1):
        for column in range(1, worksheet.max_column + 1):
            value = normalize_text(worksheet.cell(row, column).value)
            if str(year) in value and any(
                label in value for label in ("ACUMULADO", "SALDO FINAL")
            ):
                return column
    raise ValueError(f"La plantilla no contiene el acumulado del año {year} en P Y G.")


def _account_column(worksheet) -> int:
    candidates: dict[int, int] = {}
    for column in range(1, min(worksheet.max_column, 8) + 1):
        count = 0
        for row in range(1, worksheet.max_row + 1):
            value = worksheet.cell(row, column).value
            code = str(value).replace(".0", "").strip() if value is not None else ""
            if code.isdigit() and len(code) == 4:
                count += 1
        candidates[column] = count
    column, count = max(candidates.items(), key=lambda item: item[1])
    if count == 0:
        raise ValueError(
            f"No se encontró la columna de cuentas de 4 dígitos en la hoja {worksheet.title}."
        )
    return column


def _account_code(value: object) -> str:
    code = str(value).replace(".0", "").strip() if value is not None else ""
    return code if code.isdigit() and len(code) == 4 else ""


def _find_result_account_code(worksheet) -> str:
    account_column = _account_column(worksheet)
    result_labels = (
        "PERDIDAS Y GANANCIAS",
        "RESULTADO DEL EJERCICIO",
        "UTILIDAD DEL EJERCICIO",
        "PERDIDA DEL EJERCICIO",
    )
    for row in range(1, worksheet.max_row + 1):
        code = _account_code(worksheet.cell(row, account_column).value)
        label = normalize_text(worksheet.cell(row, account_column + 1).value)
        if code.startswith("3") and any(text in label for text in result_labels):
            return code
    existing_codes = {
        _account_code(worksheet.cell(row, account_column).value)
        for row in range(1, worksheet.max_row + 1)
    }
    for preferred_code in ("3605", "3610", "3705"):
        if preferred_code in existing_codes:
            return preferred_code
    raise ValueError(
        "La hoja BCE no contiene una cuenta identificable para el resultado del ejercicio."
    )


def _account_catalog(detail: pd.DataFrame) -> dict[str, str]:
    data = detail.copy()
    data["CUENTA_4"] = (
        data["CODIGO_CUENTA"]
        .fillna("")
        .astype(str)
        .str.replace(r"\.0$", "", regex=True)
        .str.replace(r"\D", "", regex=True)
        .str[:4]
    )
    name_column = next(
        (
            column
            for column in ("NOMBRE_CUENTA", "DESCRIPCION_CUENTA", "CUENTA")
            if column in data.columns
        ),
        None,
    )
    catalog: dict[str, str] = {}
    for _, row in data[data["CUENTA_4"].str.len() == 4].iterrows():
        code = str(row["CUENTA_4"])
        name = str(row.get(name_column, "") or "").strip() if name_column else ""
        catalog.setdefault(code, name or f"Cuenta {code}")
    return catalog


def _is_result_source_account(code: str, name: str) -> bool:
    label = normalize_text(name)
    return code.startswith("3") and any(
        text in label
        for text in (
            "PERDIDA DEL EJERCICIO",
            "UTILIDAD DEL EJERCICIO",
            "RESULTADO DEL EJERCICIO",
            "PERDIDAS Y GANANCIAS",
        )
    )


def _insert_result_source_account(
    worksheet,
    code: str,
    name: str,
) -> dict[str, object]:
    account_column = _account_column(worksheet)
    result_code = _find_result_account_code(worksheet)
    result_row = next(
        row
        for row in range(1, worksheet.max_row + 1)
        if _account_code(worksheet.cell(row, account_column).value) == result_code
    )
    target_row = result_row + 1
    while target_row <= worksheet.max_row:
        existing_code = _account_code(worksheet.cell(target_row, account_column).value)
        existing_label = worksheet.cell(target_row, account_column + 1).value
        if not existing_code and existing_label in (None, ""):
            break
        if not existing_code and existing_label not in (None, ""):
            raise ValueError(
                f"No hay espacio disponible junto al resultado del ejercicio para agregar {code} - {name}."
            )
        target_row += 1

    for column in range(1, worksheet.max_column + 1):
        source = worksheet.cell(result_row, column)
        target = worksheet.cell(target_row, column)
        if isinstance(target, MergedCell):
            continue
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format
        target.alignment = copy(source.alignment)
        target.border = copy(source.border)
        target.fill = copy(source.fill)
        target.font = copy(source.font)
        target.value = None
    worksheet.cell(target_row, account_column).value = code
    worksheet.cell(target_row, account_column + 1).value = name

    for row in range(1, worksheet.max_row + 1):
        label = normalize_text(worksheet.cell(row, account_column + 1).value)
        if "PATRIMONIO NETO" not in label:
            continue
        for cell in worksheet[row]:
            if not isinstance(cell.value, str) or not cell.value.startswith("="):
                continue
            reference = f"{get_column_letter(cell.column)}{target_row}"
            if reference not in cell.value.replace("$", ""):
                cell.value = f"{cell.value}-{reference}"
        break

    return {
        "Cuenta": code,
        "Nombre": name,
        "Hoja": worksheet.title,
        "Ubicación": f"Fila {target_row}",
        "Motivo": "Saldo patrimonial del resultado del ejercicio presente en el balance de prueba.",
    }


def _detail_region(worksheet, anchor_row: int) -> tuple[int, int]:
    account_column = _account_column(worksheet)
    start = anchor_row
    while start > 1:
        previous_code = _account_code(worksheet.cell(start - 1, account_column).value)
        previous_label = worksheet.cell(start - 1, account_column + 1).value
        if not previous_code and previous_label not in (None, ""):
            break
        start -= 1

    end = anchor_row
    while end < worksheet.max_row:
        next_code = _account_code(worksheet.cell(end + 1, account_column).value)
        next_label = worksheet.cell(end + 1, account_column + 1).value
        if not next_code and next_label not in (None, ""):
            break
        end += 1
    return start, end


def _expand_subtotal_formula(
    worksheet,
    subtotal_row: int,
    old_end: int,
    new_end: int,
) -> None:
    if new_end <= old_end or subtotal_row < 1:
        return
    for cell in worksheet[subtotal_row]:
        if isinstance(cell.value, str) and cell.value.startswith("="):
            cell.value = re.sub(
                rf"(?P<column>\$?[A-Z]+)\$?{old_end}\b",
                rf"\g<column>{new_end}",
                cell.value,
            )


def _ensure_statement_accounts(
    worksheet,
    catalog: dict[str, str],
    allowed_classes: set[str],
) -> list[dict[str, object]]:
    account_column = _account_column(worksheet)
    existing_rows = {
        _account_code(worksheet.cell(row, account_column).value): row
        for row in range(1, worksheet.max_row + 1)
        if _account_code(worksheet.cell(row, account_column).value)
    }
    additions: list[dict[str, object]] = []

    for code in sorted(catalog):
        name = catalog[code]
        if code[0] not in allowed_classes or code in existing_rows:
            continue
        if _is_result_source_account(code, name):
            additions.append(_insert_result_source_account(worksheet, code, name))
            existing_rows[code] = next(
                row
                for row in range(1, worksheet.max_row + 1)
                if _account_code(worksheet.cell(row, account_column).value) == code
            )
            continue

        comparable = [
            (existing_code, row)
            for existing_code, row in existing_rows.items()
            if existing_code[0] == code[0]
        ]
        same_group = [
            item for item in comparable if item[0][:2] == code[:2]
        ]
        candidates = same_group or comparable
        if not candidates:
            raise ValueError(
                f"No fue posible ubicar contablemente la cuenta {code} - {name}."
            )
        _, anchor_row = min(
            candidates,
            key=lambda item: abs(int(item[0]) - int(code)),
        )
        region_start, region_end = _detail_region(worksheet, anchor_row)

        account_rows = []
        old_last_account_row = region_start
        for row in range(region_start, region_end + 1):
            existing_code = _account_code(worksheet.cell(row, account_column).value)
            if not existing_code:
                continue
            old_last_account_row = max(old_last_account_row, row)
            account_rows.append(
                (
                    existing_code,
                    [worksheet.cell(row, column).value for column in range(1, worksheet.max_column + 1)],
                )
            )
        new_values = [None] * worksheet.max_column
        new_values[account_column - 1] = code
        new_values[account_column] = name
        account_rows.append((code, new_values))
        account_rows.sort(key=lambda item: int(item[0]))

        if len(account_rows) > region_end - region_start + 1:
            raise ValueError(
                f"No hay espacio disponible en la plantilla para agregar {code} - {name}."
            )
        for offset, row in enumerate(range(region_start, region_end + 1)):
            values = account_rows[offset][1] if offset < len(account_rows) else [None] * worksheet.max_column
            for column, value in enumerate(values, start=1):
                cell = worksheet.cell(row, column)
                if not isinstance(cell, MergedCell):
                    cell.value = value

        _expand_subtotal_formula(
            worksheet,
            region_start - 1,
            old_last_account_row,
            region_start + len(account_rows) - 1,
        )
        existing_rows = {
            _account_code(worksheet.cell(row, account_column).value): row
            for row in range(1, worksheet.max_row + 1)
            if _account_code(worksheet.cell(row, account_column).value)
        }
        additions.append(
            {
                "Cuenta": code,
                "Nombre": name,
                "Hoja": worksheet.title,
                "Ubicación": f"Fila {existing_rows[code]}",
                "Motivo": "Cuenta presente en el balance de prueba y ausente en la plantilla.",
            }
        )
    return additions


def _write_mapped_accounts(worksheet, column: int, values: dict[str, float]) -> int:
    updated = 0
    account_column = _account_column(worksheet)
    for row in range(1, worksheet.max_row + 1):
        raw_code = worksheet.cell(row, account_column).value
        if raw_code is None:
            continue
        code = str(raw_code).replace(".0", "").strip()
        if not code.isdigit() or len(code) != 4:
            continue
        target_cell = worksheet.cell(row, column)
        if isinstance(target_cell, MergedCell):
            continue
        target_cell.value = float(values.get(code, 0.0))
        updated += 1
    return updated


def _replace_external_formulas(workbook, cached_workbook) -> None:
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        cached_sheet = cached_workbook[sheet_name]
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                if (
                    isinstance(cell.value, str)
                    and cell.value.startswith("=")
                    and "[" in cell.value
                ):
                    cached_value = cached_sheet[cell.coordinate].value
                    cell.value = cached_value if cached_value is not None else 0
                elif isinstance(cell.value, str) and "#REF!" in cell.value:
                    cached_value = cached_sheet[cell.coordinate].value
                    cell.value = (
                        cached_value
                        if cached_value not in (None, "#REF!")
                        else 0
                    )
    workbook._external_links = []


def _configure_visible_periods(
    workbook,
    start_year: int,
    monthly_periods: list[tuple[int, int]],
    preserve_existing: bool = False,
) -> None:
    bce = workbook["BCE"]
    pyg = workbook["P Y G"]

    base_bce_column = _find_bce_month_column(bce, start_year, 12)
    pyg_base_column = _find_pyg_total_column(pyg, start_year)
    for year, month in monthly_periods:
        bce_column = _find_bce_month_column(bce, year, month)
        _find_pyg_month_column(pyg, bce_column, year, month)

    bce_header = bce.cell(2, base_bce_column)
    if not isinstance(bce_header, MergedCell):
        bce_header.value = f"Saldo final {start_year}"
    pyg_header = pyg.cell(3, pyg_base_column)
    if not isinstance(pyg_header, MergedCell):
        pyg_header.value = f"Saldo final {start_year}"

    bce.sheet_state = "visible"
    pyg.sheet_state = "visible"
    workbook.active = workbook.index(bce)
    for worksheet in workbook.worksheets:
        worksheet.sheet_view.tabSelected = worksheet is bce


def _make_workbook_fully_editable(workbook) -> None:
    if getattr(workbook, "security", None) is not None:
        workbook.security.lockStructure = False
        workbook.security.lockWindows = False

    active_worksheet = workbook.active
    for worksheet in workbook.worksheets:
        worksheet.sheet_state = "visible"
        worksheet.sheet_view.tabSelected = worksheet is active_worksheet
        worksheet.protection.sheet = False
        worksheet.protection.enable()
        worksheet.protection.disable()
        worksheet.auto_filter.ref = None
        worksheet.sheet_view.view = "normal"

        for dimension in worksheet.column_dimensions.values():
            dimension.hidden = False
            dimension.collapsed = False
            dimension.outlineLevel = 0
        for dimension in worksheet.row_dimensions.values():
            dimension.hidden = False
            dimension.collapsed = False
            dimension.outlineLevel = 0
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                cell.protection = Protection(locked=False, hidden=False)


def _write_validation_sheet(
    workbook,
    additions: list[dict[str, object]],
    metric_rows: list[dict[str, object]],
) -> None:
    if "Validación" in workbook.sheetnames:
        del workbook["Validación"]
    worksheet = workbook.create_sheet("Validación")
    worksheet.sheet_view.showGridLines = False
    worksheet["A1"] = "Validación contable del informe mensualizado"
    worksheet["A1"].font = Font(bold=True, size=15, color="FFFFFF")
    worksheet["A1"].fill = PatternFill("solid", fgColor="0B6B57")
    worksheet.merge_cells("A1:F1")

    worksheet.append([])
    worksheet.append(
        ["Periodo", "Diferencia de cuadre", "Estado", "Activo", "Pasivo", "Patrimonio"]
    )
    for metric in metric_rows:
        difference = float(metric["Diferencia de cuadre"])
        worksheet.append(
            [
                metric["Periodo"],
                difference,
                "CUADRA" if abs(difference) <= 1 else "REVISAR",
                float(metric["Activos"]),
                float(metric["Pasivos"]),
                float(metric["Patrimonio"]),
            ]
        )

    start_row = worksheet.max_row + 3
    worksheet.cell(start_row, 1).value = "Cuentas agregadas automáticamente"
    worksheet.cell(start_row, 1).font = Font(bold=True, color="FFFFFF")
    worksheet.cell(start_row, 1).fill = PatternFill("solid", fgColor="344054")
    headers = ["Cuenta", "Nombre", "Hoja", "Ubicación", "Motivo"]
    for column, header in enumerate(headers, start=1):
        worksheet.cell(start_row + 1, column).value = header
    for addition in additions:
        worksheet.append([addition.get(header, "") for header in headers])
    if not additions:
        worksheet.append(["-", "No fue necesario crear cuentas nuevas.", "-", "-", "-"])

    for row in (3, start_row + 1):
        for cell in worksheet[row]:
            if cell.value is not None:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F2937")
    for column in ("B", "D", "E", "F"):
        for cell in worksheet[column]:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '$#,##0;[Red]($#,##0);-'
    widths = {"A": 18, "B": 42, "C": 16, "D": 18, "E": 58, "F": 18}
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width


def _load_template(previous_file: BinaryIO | None):
    if previous_file is not None:
        previous_file.seek(0)
        source_bytes = previous_file.read()
        source = BytesIO(source_bytes)
        cached_source = BytesIO(source_bytes)
        source_name = "Informe mensualizado anterior"
    else:
        if not TEMPLATE_PATH.exists():
            raise FileNotFoundError("No se encontró la plantilla oficial mensualizada.")
        source = TEMPLATE_PATH
        cached_source = TEMPLATE_PATH
        source_name = "Plantilla oficial"

    workbook = load_workbook(source, data_only=False, keep_links=False)
    cached_workbook = load_workbook(cached_source, data_only=True, keep_links=False)
    missing = {"BCE", "P Y G"} - set(workbook.sheetnames)
    if missing:
        raise ValueError("La plantilla debe contener las hojas BCE y P Y G.")
    _replace_external_formulas(workbook, cached_workbook)
    return workbook, source_name


def _infer_start_year(workbook) -> int:
    bce = workbook["BCE"]
    for column in range(1, bce.max_column + 1):
        value = bce.cell(2, column).value
        if isinstance(value, str) and value.lower().startswith("saldo final"):
            year_text = "".join(character for character in value if character.isdigit())
            if len(year_text) == 4:
                return int(year_text)
    for year in sorted(SUPPORTED_YEARS, reverse=True):
        try:
            column = _find_bce_month_column(bce, year, 12)
        except ValueError:
            continue
        if not bce.column_dimensions[get_column_letter(column)].hidden:
            return year
    return 2025


def _period_metrics(
    detail: pd.DataFrame,
    year: int,
    month: int,
    label: str,
    kind: str,
) -> dict[str, object]:
    codes = detail["CODIGO_CUENTA"].astype(str)

    def movement(prefixes: tuple[str, ...], income_sign: bool = False) -> float:
        selected = detail[codes.str.startswith(prefixes)]
        if income_sign:
            return float(
                selected["MOVIMIENTO_CREDITO"].sum()
                - selected["MOVIMIENTO_DEBITO"].sum()
            )
        return float(
            selected["MOVIMIENTO_DEBITO"].sum()
            - selected["MOVIMIENTO_CREDITO"].sum()
        )

    operating_income = movement(("41",), income_sign=True)
    other_income = movement(("42", "47"), income_sign=True)
    direct_costs = movement(("61", "62", "71", "72", "73"))
    administration = movement(("51",))
    selling = movement(("52",))
    other_expenses = movement(("53", "54"))
    income = operating_income + other_income
    costs = direct_costs + administration + selling + other_expenses
    gross_profit = operating_income - direct_costs
    operating_profit = gross_profit - administration - selling
    net_result = income - costs
    depreciation = movement(("5160", "5165", "5260"))
    assets = detail.loc[detail["CLASE"] == "1", "SALDO_FINAL"].sum()
    liabilities = abs(detail.loc[detail["CLASE"] == "2", "SALDO_FINAL"].sum())
    equity = abs(detail.loc[detail["CLASE"] == "3", "SALDO_FINAL"].sum())
    accumulated_result = detail.loc[
        detail["CLASE"].isin(["4", "5", "6", "7"]), "SALDO_FINAL"
    ].sum()
    balance_difference = assets - (liabilities + equity - accumulated_result)
    current_assets = detail.loc[
        codes.str.startswith(("11", "13", "14")), "SALDO_FINAL"
    ].sum()
    current_liabilities = abs(
        detail.loc[codes.str.startswith(("21", "22", "23", "24", "25", "26", "27", "28")), "SALDO_FINAL"].sum()
    )
    return {
        "Tipo": kind,
        "Periodo": label,
        "Año": year,
        "Mes": month,
        "Ingresos": float(income),
        "Ventas": float(operating_income),
        "Utilidad bruta": float(gross_profit),
        "Utilidad operacional": float(operating_profit),
        "Costos y gastos": float(costs),
        "Resultado": float(net_result),
        "EBITDA": float(operating_profit + depreciation),
        "Activos": float(assets),
        "Pasivos": float(liabilities),
        "Patrimonio": float(equity),
        "Activos corrientes": float(current_assets),
        "Pasivos corrientes": float(current_liabilities),
        "Resultado acumulado": float(accumulated_result),
        "Diferencia de cuadre": float(balance_difference),
    }


def _transactional_rows(raw_df: pd.DataFrame) -> pd.DataFrame:
    transactional = raw_df["TRANSACCIONAL"].map(normalize_text).eq("SI")
    detail = raw_df[transactional].copy()
    if detail.empty:
        max_length = raw_df["CODIGO_CUENTA"].astype(str).str.len().max()
        detail = raw_df[
            raw_df["CODIGO_CUENTA"].astype(str).str.len().eq(max_length)
        ].copy()
    return detail


def _third_party_tables(raw_df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    detail = _transactional_rows(raw_df)
    detail["CUENTA"] = detail["CODIGO_CUENTA"].astype(str)
    detail["Tercero"] = (
        detail["NOMBRE_TERCERO"].fillna("").astype(str).str.strip()
    )
    detail.loc[detail["Tercero"].eq(""), "Tercero"] = "Sin tercero"
    detail["Identificación"] = (
        detail["IDENTIFICACION"].fillna("").astype(str).str.replace(r"\.0$", "", regex=True)
    )

    receivables = detail[detail["CUENTA"].str.startswith("13")].copy()
    receivables = (
        receivables.groupby(["Identificación", "Tercero"], as_index=False)["SALDO_FINAL"]
        .sum()
        .rename(columns={"SALDO_FINAL": "Saldo"})
    )
    receivables["Saldo"] = receivables["Saldo"].clip(lower=0)
    receivables = receivables.sort_values("Saldo", ascending=False)
    receivables = receivables[receivables["Saldo"].abs() > 0.5]

    payables = detail[
        detail["CUENTA"].str.startswith(("21", "22", "23", "24", "25", "26", "27", "28"))
    ].copy()
    payables = (
        payables.groupby(["Identificación", "Tercero"], as_index=False)["SALDO_FINAL"]
        .sum()
        .rename(columns={"SALDO_FINAL": "Saldo"})
    )
    payables["Saldo"] = payables["Saldo"].abs()
    payables = payables.sort_values("Saldo", ascending=False)
    payables = payables[payables["Saldo"].abs() > 0.5]

    activity = detail.copy()
    activity["Movimiento"] = (
        activity["MOVIMIENTO_DEBITO"].abs() + activity["MOVIMIENTO_CREDITO"].abs()
    )
    activity = (
        activity.groupby(["Identificación", "Tercero"], as_index=False)
        .agg(
            Débitos=("MOVIMIENTO_DEBITO", "sum"),
            Créditos=("MOVIMIENTO_CREDITO", "sum"),
            Movimiento=("Movimiento", "sum"),
        )
        .sort_values("Movimiento", ascending=False)
    )
    activity = activity[activity["Movimiento"].abs() > 0.5]
    return {
        "receivables": receivables.reset_index(drop=True),
        "payables": payables.reset_index(drop=True),
        "activity": activity.reset_index(drop=True),
    }


def _expense_composition(detail: pd.DataFrame) -> pd.DataFrame:
    data = detail.copy()
    data["Grupo"] = data["CODIGO_CUENTA"].astype(str).str[:2]
    labels = {
        "51": "Administración",
        "52": "Ventas",
        "53": "No operacionales",
        "54": "Impuesto de renta",
        "61": "Costo de ventas",
        "62": "Costo de ventas",
        "71": "Costos de producción",
        "72": "Costos de producción",
        "73": "Costos de producción",
    }
    data = data[data["Grupo"].isin(labels)].copy()
    data["Concepto"] = data["Grupo"].map(labels)
    data["Valor"] = data["MOVIMIENTO_DEBITO"] - data["MOVIMIENTO_CREDITO"]
    return (
        data.groupby("Concepto", as_index=False)["Valor"]
        .sum()
        .sort_values("Valor", ascending=False)
        .reset_index(drop=True)
    )


def _statement_account_labels(worksheet) -> dict[str, str]:
    account_column = _account_column(worksheet)
    labels: dict[str, str] = {}
    for row in range(1, worksheet.max_row + 1):
        code = _account_code(worksheet.cell(row, account_column).value)
        if code:
            labels[code] = (
                str(worksheet.cell(row, account_column + 1).value or f"Cuenta {code}")
                .strip()
            )
    return labels


def _client_period_view(
    raw_df: pd.DataFrame,
    period_label: str,
    balance_values: dict[str, float],
    pyg_values: dict[str, float],
    result_account_code: str,
    balance_labels: dict[str, str],
    pyg_labels: dict[str, str],
) -> dict[str, object]:
    summary_rows: list[dict[str, object]] = []
    ordered_balance_codes = list(balance_labels)
    ordered_balance_codes.extend(
        code for code in balance_values if code not in balance_labels
    )
    for order, code in enumerate(ordered_balance_codes):
        value = balance_values.get(code, 0.0)
        if code[:1] not in {"1", "2", "3"}:
            continue
        summary_rows.append(
            {
                "Estado": "balance",
                "Codigo": code,
                "Concepto": balance_labels.get(code, f"Cuenta {code}"),
                "Orden": order,
                "Periodo": period_label,
                "Valor": abs(float(value)),
                "ValorCalculo": float(value),
            }
        )

    ordered_pyg_codes = list(pyg_labels)
    ordered_pyg_codes.extend(
        code for code in pyg_values if code not in pyg_labels
    )
    for order, code in enumerate(ordered_pyg_codes):
        value = pyg_values.get(code, 0.0)
        if code[:1] not in {"4", "5", "6", "7"}:
            continue
        summary_rows.append(
            {
                "Estado": "pyg",
                "Codigo": code,
                "Concepto": pyg_labels.get(code, f"Cuenta {code}"),
                "Orden": order,
                "Periodo": period_label,
                "Valor": float(value),
                "ValorCalculo": float(value),
            }
        )

    detail = _transactional_rows(raw_df).copy()
    detail["CUENTA_4"] = (
        detail["CODIGO_CUENTA"]
        .fillna("")
        .astype(str)
        .str.replace(r"\.0$", "", regex=True)
        .str.replace(r"\D", "", regex=True)
        .str[:4]
    )
    full_codes = (
        detail["CODIGO_CUENTA"]
        .fillna("")
        .astype(str)
        .str.replace(r"\.0$", "", regex=True)
        .str.replace(r"\D", "", regex=True)
    )
    detail = detail[full_codes.str.len().eq(8)].copy()
    detail["CUENTA_4"] = full_codes.loc[detail.index].str[:4]
    detail["Identificación"] = (
        detail["IDENTIFICACION"]
        .fillna("")
        .astype(str)
        .str.replace(r"\.0$", "", regex=True)
        .str.strip()
    )
    detail["Tercero"] = detail["NOMBRE_TERCERO"].fillna("").astype(str).str.strip()
    detail.loc[detail["Tercero"].eq(""), "Tercero"] = "Sin tercero identificado"

    balance_detail = detail[detail["CLASE"].isin(["1", "2", "3"])].copy()
    balance_detail = balance_detail[balance_detail["CUENTA_4"] != result_account_code]
    balance_detail["Saldo"] = balance_detail["SALDO_FINAL"]
    balance_detail = (
        balance_detail.groupby(
            ["CUENTA_4", "Identificación", "Tercero"],
            as_index=False,
        )["Saldo"]
        .sum()
    )
    synthetic_result = float(balance_values.get(result_account_code, 0.0))
    if abs(synthetic_result) > 0.5:
        balance_detail = pd.concat(
            [
                balance_detail,
                pd.DataFrame(
                    [
                        {
                            "CUENTA_4": result_account_code,
                            "Identificación": "-",
                            "Tercero": "Resultado acumulado del Estado de Resultados",
                            "Saldo": synthetic_result,
                        }
                    ]
                ),
            ],
            ignore_index=True,
        )

    pyg_detail = detail[detail["CLASE"].isin(["4", "5", "6", "7"])].copy()
    pyg_detail["Saldo"] = (
        pyg_detail["MOVIMIENTO_DEBITO"] - pyg_detail["MOVIMIENTO_CREDITO"]
    )
    income_rows = pyg_detail["CLASE"].eq("4")
    pyg_detail.loc[income_rows, "Saldo"] = (
        pyg_detail.loc[income_rows, "MOVIMIENTO_CREDITO"]
        - pyg_detail.loc[income_rows, "MOVIMIENTO_DEBITO"]
    )
    pyg_detail = (
        pyg_detail.groupby(
            ["CUENTA_4", "Identificación", "Tercero"],
            as_index=False,
        )["Saldo"]
        .sum()
    )

    details: dict[str, pd.DataFrame] = {}
    for statement, frame in (("balance", balance_detail), ("pyg", pyg_detail)):
        labels = balance_labels if statement == "balance" else pyg_labels
        for code, grouped in frame.groupby("CUENTA_4"):
            clean = grouped[grouped["Saldo"].abs() > 0.5].copy()
            clean["Concepto"] = labels.get(str(code), f"Cuenta {code}")
            clean["Periodo"] = period_label
            clean = clean.sort_values("Saldo", key=lambda series: series.abs(), ascending=False)
            details[f"{statement}|{period_label}|{code}"] = clean[
                ["Periodo", "Concepto", "Identificación", "Tercero", "Saldo"]
            ].reset_index(drop=True)
    return {
        "summary": summary_rows,
        "details": details,
    }


def _finalize_client_financial_view(
    summary_rows: list[dict[str, object]],
    details: dict[str, pd.DataFrame],
    opening_period: str | None,
    monthly_periods: list[tuple[int, int]],
) -> dict[str, object]:
    summary = pd.DataFrame(
        summary_rows,
        columns=[
            "Estado",
            "Codigo",
            "Concepto",
            "Orden",
            "Periodo",
            "Valor",
            "ValorCalculo",
        ],
    )
    if not summary.empty:
        summary["Orden"] = summary.groupby(
            ["Estado", "Codigo"],
        )["Orden"].transform("min")
    period_order: list[str] = [opening_period] if opening_period else []

    for year in sorted({year for year, _ in monthly_periods}):
        year_periods = [
            f"{month:02d}/{period_year}"
            for period_year, month in monthly_periods
            if period_year == year
        ]
        period_order.extend(year_periods)
        accumulated_label = f"Acumulado {year}"
        pyg_year = summary[
            summary["Estado"].eq("pyg") & summary["Periodo"].isin(year_periods)
        ]
        if pyg_year.empty:
            continue
        accumulated = (
            pyg_year.groupby(
                ["Estado", "Codigo", "Concepto", "Orden"],
                as_index=False,
            )[["Valor", "ValorCalculo"]]
            .sum()
        )
        accumulated["Periodo"] = accumulated_label
        summary = pd.concat([summary, accumulated], ignore_index=True)
        period_order.append(accumulated_label)

        for code in accumulated["Codigo"].astype(str).unique():
            frames = [
                details[key]
                for period in year_periods
                if (key := f"pyg|{period}|{code}") in details
            ]
            if not frames:
                continue
            combined = pd.concat(frames, ignore_index=True)
            combined = (
                combined.groupby(
                    ["Concepto", "Identificación", "Tercero"],
                    as_index=False,
                )["Saldo"]
                .sum()
            )
            combined = combined[combined["Saldo"].abs() > 0.5]
            combined["Periodo"] = accumulated_label
            details[f"pyg|{accumulated_label}|{code}"] = combined[
                ["Periodo", "Concepto", "Identificación", "Tercero", "Saldo"]
            ].sort_values(
                "Saldo",
                key=lambda series: series.abs(),
                ascending=False,
            ).reset_index(drop=True)

    return {
        "summary": summary,
        "details": details,
        "period_order": period_order,
    }


def _statement_preview(
    worksheet,
    values: dict[str, float],
    value_label: str,
) -> pd.DataFrame:
    rows = []
    account_column = _account_column(worksheet)
    concept_column = account_column + 1
    for row in range(1, worksheet.max_row + 1):
        raw_code = worksheet.cell(row, account_column).value
        if raw_code is None:
            continue
        code = str(raw_code).replace(".0", "").strip()
        if not code.isdigit() or len(code) != 4:
            continue
        rows.append(
            {
                "Cuenta": code,
                "Concepto": worksheet.cell(row, concept_column).value or "Sin descripción",
                value_label: float(values.get(code, 0.0)),
            }
        )
    return pd.DataFrame(rows, columns=["Cuenta", "Concepto", value_label])


def build_monthly_reports(
    monthly_files: list[BinaryIO],
    previous_file: BinaryIO | None = None,
    initial_balance_file: BinaryIO | None = None,
    start_year: int | None = None,
) -> dict[str, object]:
    if not monthly_files:
        raise ValueError("Debes cargar al menos un balance mensual.")
    if previous_file is None and initial_balance_file is None:
        raise ValueError("Debes cargar el saldo inicial o un informe anterior.")

    workbook, source_name = _load_template(previous_file)
    bce = workbook["BCE"]
    pyg = workbook["P Y G"]
    summary_rows: list[dict[str, object]] = []
    metric_rows: list[dict[str, object]] = []
    monthly_periods: list[tuple[int, int]] = []
    latest_balance_values: dict[str, float] = {}
    latest_pyg_values: dict[str, float] = {}
    latest_raw_df = pd.DataFrame()
    latest_detail = pd.DataFrame()
    account_additions: list[dict[str, object]] = []
    client_summary_rows: list[dict[str, object]] = []
    client_details: dict[str, pd.DataFrame] = {}

    if initial_balance_file is not None:
        initial_balance_file.seek(0)
        opening_df, opening_metadata = load_tlg_trial_balance(initial_balance_file)
        opening_detail = prepare_tlg_detail(opening_df)
        opening_year = int(start_year or opening_metadata.get("anio") or 0)
        if opening_year not in SUPPORTED_YEARS:
            raise ValueError(
                f"La plantilla actual solo contiene los años {', '.join(map(str, sorted(SUPPORTED_YEARS)))}."
            )
        opening_catalog = _account_catalog(opening_detail)
        account_additions.extend(
            _ensure_statement_accounts(bce, opening_catalog, {"1", "2", "3"})
        )
        account_additions.extend(
            _ensure_statement_accounts(pyg, opening_catalog, {"4", "5", "6", "7"})
        )
        result_account_code = _find_result_account_code(bce)
        opening_balance, opening_pyg = _account4_values(
            opening_detail,
            result_account_code,
        )
        opening_view = _client_period_view(
            opening_df,
            f"Saldo final {opening_year}",
            opening_balance,
            opening_pyg,
            result_account_code,
            _statement_account_labels(bce),
            _statement_account_labels(pyg),
        )
        client_summary_rows.extend(opening_view["summary"])
        client_details.update(opening_view["details"])
        bce_base_column = _find_bce_month_column(bce, opening_year, 12)
        pyg_base_column = _find_pyg_total_column(pyg, opening_year)
        bce_count = _write_mapped_accounts(bce, bce_base_column, opening_balance)
        pyg_count = _write_mapped_accounts(pyg, pyg_base_column, opening_pyg)
        summary_rows.append(
            {
                "Tipo": "Saldo inicial",
                "Periodo": f"Saldo final {opening_year}",
                "Archivo": getattr(initial_balance_file, "name", "saldo_inicial.xlsx"),
                "Cuentas leídas": int(opening_detail["CODIGO_CUENTA"].nunique()),
                "Filas BCE actualizadas": bce_count,
                "Filas PYG actualizadas": pyg_count,
            }
        )
        metric_rows.append(
            _period_metrics(
                opening_detail,
                opening_year,
                0,
                f"Saldo final {opening_year}",
                "Saldo inicial",
            )
        )
    else:
        opening_year = int(start_year or _infer_start_year(workbook))
        result_account_code = _find_result_account_code(bce)

    seen_periods: set[tuple[int, int]] = set()
    for uploaded_file in monthly_files:
        uploaded_file.seek(0)
        raw_df, metadata = load_tlg_trial_balance(uploaded_file)
        year, month = _period_from_metadata(metadata)
        if (year, month) in seen_periods:
            raise ValueError(f"Se cargó más de un archivo para {month:02d}/{year}.")
        seen_periods.add((year, month))
        detail = prepare_tlg_detail(raw_df)
        latest_raw_df = raw_df
        latest_detail = detail
        catalog = _account_catalog(detail)
        account_additions.extend(
            _ensure_statement_accounts(bce, catalog, {"1", "2", "3"})
        )
        account_additions.extend(
            _ensure_statement_accounts(pyg, catalog, {"4", "5", "6", "7"})
        )
        result_account_code = _find_result_account_code(bce)
        balance_values, pyg_values = _account4_values(
            detail,
            result_account_code,
        )
        period_label = f"{month:02d}/{year}"
        period_view = _client_period_view(
            raw_df,
            period_label,
            balance_values,
            pyg_values,
            result_account_code,
            _statement_account_labels(bce),
            _statement_account_labels(pyg),
        )
        client_summary_rows.extend(period_view["summary"])
        client_details.update(period_view["details"])
        latest_balance_values = balance_values
        latest_pyg_values = pyg_values
        bce_column = _find_bce_month_column(bce, year, month)
        pyg_column = _find_pyg_month_column(pyg, bce_column, year, month)
        bce_count = _write_mapped_accounts(bce, bce_column, balance_values)
        pyg_count = _write_mapped_accounts(pyg, pyg_column, pyg_values)
        monthly_periods.append((year, month))
        summary_rows.append(
            {
                "Tipo": "Mensual",
                "Periodo": f"{month:02d}/{year}",
                "Archivo": getattr(uploaded_file, "name", f"{month:02d}-{year}.xlsx"),
                "Cuentas leídas": int(detail["CODIGO_CUENTA"].nunique()),
                "Filas BCE actualizadas": bce_count,
                "Filas PYG actualizadas": pyg_count,
            }
        )
        metric_rows.append(
            _period_metrics(detail, year, month, f"{month:02d}/{year}", "Mensual")
        )

    monthly_periods.sort()
    _configure_visible_periods(
        workbook,
        opening_year,
        monthly_periods,
        preserve_existing=previous_file is not None,
    )
    unique_additions = list(
        {
            (str(item["Cuenta"]), str(item["Hoja"])): item
            for item in account_additions
        }.values()
    )
    _write_validation_sheet(workbook, unique_additions, metric_rows)
    _make_workbook_fully_editable(workbook)

    if hasattr(workbook, "calculation"):
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"

    output = BytesIO()
    workbook.save(output)
    last_year, last_month = monthly_periods[-1]
    metrics = pd.DataFrame(metric_rows)
    third_parties = _third_party_tables(latest_raw_df)
    expense_composition = _expense_composition(latest_detail)
    client_financial_view = _finalize_client_financial_view(
        client_summary_rows,
        client_details,
        f"Saldo final {opening_year}" if initial_balance_file is not None else None,
        monthly_periods,
    )
    balance_preview = _statement_preview(
        bce,
        latest_balance_values,
        f"Saldo {last_month:02d}/{last_year}",
    )
    pyg_preview = _statement_preview(
        pyg,
        latest_pyg_values,
        f"Movimiento {last_month:02d}/{last_year}",
    )
    return {
        "output": output.getvalue(),
        "periods": pd.DataFrame(summary_rows),
        "metrics": metrics,
        "balance_preview": balance_preview,
        "pyg_preview": pyg_preview,
        "third_party_receivables": third_parties["receivables"],
        "third_party_payables": third_parties["payables"],
        "third_party_activity": third_parties["activity"],
        "expense_composition": expense_composition,
        "account_additions": pd.DataFrame(unique_additions) if unique_additions else pd.DataFrame(
            columns=["Cuenta", "Nombre", "Hoja", "Ubicación", "Motivo"]
        ),
        "client_financial_view": client_financial_view,
        "source_name": source_name,
        "start_year": opening_year,
        "last_period": f"{last_month:02d}/{last_year}",
    }


def export_monthly_reports(report: dict[str, object]) -> bytes:
    return bytes(report["output"])
